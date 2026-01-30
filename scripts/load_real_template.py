#!/usr/bin/env python3
"""
Load real inspection template from Excel file.
Designed for Power Park Student Housing Phase 3.
Uses openpyxl (no pandas dependency).
"""

import sqlite3
import uuid
from openpyxl import load_workbook

# Configuration
DB_PATH = '/var/data/inspections.db'
EXCEL_PATH = 'attached_assets/Defective_works_Unit_empty_20260126.xlsx'
TENANT_ID = 'MONOGRAPH'
UNIT_TYPE = 'STANDARD'

# Area definitions: (start_row, end_row, area_name) - 1-indexed for openpyxl
AREAS = [
    (24, 200, 'KITCHEN'),
    (201, 241, 'LOUNGE'),
    (242, 355, 'BATHROOM'),
    (356, 446, 'BEDROOM A'),
    (447, 537, 'BEDROOM B'),
    (538, 628, 'BEDROOM C'),
    (629, 748, 'BEDROOM D'),
]


def generate_id():
    return str(uuid.uuid4())


def get_cell_value(ws, row, col):
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def is_category_header(col0, col1):
    if not col0:
        return False
    if col1 in ['False', 'True', 'FALSE', 'TRUE']:
        return False
    if col0.lower() == 'description' or col0 == 'General':
        return False
    if 'AREA' in col0.upper():
        return False
    return col0.isupper()


def is_inspection_item(col0, col1):
    return col0 and col1 in ['False', 'True', 'FALSE', 'TRUE']


def load_template():
    print("Loading Excel file...")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}")
    
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    print("Clearing existing template data...")
    cur.execute("DELETE FROM item_template WHERE tenant_id = ?", (TENANT_ID,))
    cur.execute("DELETE FROM area_template WHERE tenant_id = ?", (TENANT_ID,))
    conn.commit()
    
    print("Inserting areas...")
    area_ids = {}
    for area_order, (start, end, area_name) in enumerate(AREAS, start=1):
        area_id = generate_id()
        area_ids[area_name] = area_id
        cur.execute("""
            INSERT INTO area_template (id, tenant_id, unit_type, area_name, area_order)
            VALUES (?, ?, ?, ?, ?)
        """, (area_id, TENANT_ID, UNIT_TYPE, area_name, area_order))
    
    print(f"  Inserted {len(AREAS)} areas")
    
    print("Inserting items...")
    total_items = 0
    
    for start, end, area_name in AREAS:
        area_id = area_ids[area_name]
        current_category_id = None
        item_order = 0
        area_item_count = 0
        
        for row in range(start, end + 1):
            col0 = get_cell_value(ws, row, 1)
            col1 = get_cell_value(ws, row, 2)
            
            if not col0:
                continue
            
            if is_category_header(col0, col1):
                current_category_id = generate_id()
                item_order += 1
                cur.execute("""
                    INSERT INTO item_template 
                    (id, tenant_id, category_id, parent_item_id, item_description, item_order, depth)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (current_category_id, TENANT_ID, area_id, None, col0, item_order, 0))
                
            elif is_inspection_item(col0, col1):
                item_id = generate_id()
                item_order += 1
                area_item_count += 1
                cur.execute("""
                    INSERT INTO item_template 
                    (id, tenant_id, category_id, parent_item_id, item_description, item_order, depth)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (item_id, TENANT_ID, area_id, current_category_id, col0, item_order, 1))
        
        print(f"  {area_name}: {area_item_count} items")
        total_items += area_item_count
    
    conn.commit()
    wb.close()
    
    cur.execute("SELECT area_name FROM area_template WHERE tenant_id = ? ORDER BY area_order", (TENANT_ID,))
    areas_result = [row[0] for row in cur.fetchall()]
    
    cur.execute("SELECT COUNT(*) FROM item_template WHERE tenant_id = ?", (TENANT_ID,))
    item_count = cur.fetchone()[0]
    
    conn.close()
    
    print()
    print(f"Done! Total items: {total_items} (in DB: {item_count})")
    print(f"Areas: {areas_result}")


if __name__ == '__main__':
    load_template()
