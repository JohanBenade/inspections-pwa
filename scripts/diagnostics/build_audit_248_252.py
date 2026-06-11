#!/usr/bin/env python3
"""
build_audit_248_252.py  -- forensic audit-trail workbook for units 248 + 252.

Reads /tmp/audit_248_252.csv (produced by the read-only extract on Render),
emits /tmp/Unit_248_252_Audit_Trail.xlsx with 5 sheets:
  248-C1, 248-C2, 252-C1, 252-C2  -- per-item forensic rows in walk order
  Reconciliation                  -- per-inspection status cross-foot to 508

Read-only w.r.t. the DB (consumes the CSV only). ASCII only. No hardcoded
per-unit data -- everything is driven by the CSV contents.
"""
import csv
from collections import OrderedDict, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/tmp/audit_248_252.csv"
OUT = "/tmp/Unit_248_252_Audit_Trail.xlsx"

SHEETS = ["248-C1", "248-C2", "252-C1", "252-C2"]

# ---- styling constants (DM Sans not embeddable in xlsx; use Aptos Narrow to
# ---- match the pilot sheets; gold + dark per house palette) -------------
GOLD = "C8963E"
DARK = "0A0A0A"
GREY = "D9D9D9"
LIGHT = "F5F3EE"
RED = "C0504D"
HDR_FONT = Font(name="Aptos Narrow", bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill("solid", fgColor=DARK)
AREA_FONT = Font(name="Aptos Narrow", bold=True, size=10, color=DARK)
AREA_FILL = PatternFill("solid", fgColor=GREY)
CELL_FONT = Font(name="Aptos Narrow", size=9)
PRIOR_FONT = Font(name="Aptos Narrow", size=9, italic=True, color=RED)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# CSV columns:
# sheet,unit,cycle,area_name,category_name,item_description,status,comment,
# marked_at,created_at,updated_at,has_prior_defects,area_order,category_order,item_order

COLS = [
    ("Area", "area_name", 22),
    ("Category", "category_name", 20),
    ("Item", "item_description", 34),
    ("Verdict", "_verdict", 10),
    ("Marked At", "marked_at", 20),
    ("Comment", "comment", 40),
    ("Prior Defect", "_prior", 11),
    ("Created At", "created_at", 20),
    ("Updated At", "updated_at", 20),
]

VERDICT = {
    "ok": "MS",
    "not_to_standard": "NTS",
    "not_installed": "NI",
    "skipped": "SKIP",
    "pending": "PENDING",
}


def load():
    by_sheet = OrderedDict((s, []) for s in SHEETS)
    with open(SRC, newline="") as f:
        for r in csv.DictReader(f):
            by_sheet.setdefault(r["sheet"], []).append(r)
    return by_sheet


def verdict(row):
    return VERDICT.get(row["status"], row["status"])


def prior(row):
    return "YES" if str(row.get("has_prior_defects", "")) in ("1", "1.0") else ""


def write_audit_sheet(ws, rows):
    # header
    for ci, (label, _, width) in enumerate(COLS, 1):
        cell = ws.cell(1, ci, label)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"
    r = 2
    cur_area = None
    for row in rows:
        # area band when area changes
        if row["area_name"] != cur_area:
            cur_area = row["area_name"]
            ws.cell(r, 1, cur_area)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
            bc = ws.cell(r, 1)
            bc.font = AREA_FONT
            bc.fill = AREA_FILL
            bc.alignment = Alignment(horizontal="left", vertical="center")
            r += 1
        for ci, (_, key, _) in enumerate(COLS, 1):
            if key == "_verdict":
                val = verdict(row)
            elif key == "_prior":
                val = prior(row)
            else:
                val = row.get(key, "")
            cell = ws.cell(r, ci, val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(key in ("item_description", "comment")))
            cell.font = PRIOR_FONT if (key == "_prior" and val == "YES") else CELL_FONT
        r += 1
    return r - 1


def write_recon(ws, by_sheet):
    title = ws.cell(1, 1, "Reconciliation -- Units 248 & 252 Audit Trail")
    title.font = Font(name="Aptos Narrow", bold=True, size=14, color=DARK)
    ws.merge_cells("A1:H1")
    sub = ws.cell(2, 1, "Per-inspection cross-foot to the 508 active-template universe. "
                        "Cert claim: 502 MS + 6 SKIP (structural ground-only), 0 open. Self-evident below.")
    sub.font = Font(name="Aptos Narrow", size=9, color="9A9A9A")
    ws.merge_cells("A2:H2")

    headers = ["Sheet", "MS (ok)", "NTS", "NI", "SKIP", "PENDING", "Total", "Prior-Defect Items"]
    widths = [12, 10, 8, 8, 8, 10, 9, 18]
    hr = 4
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(hr, ci, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    r = hr + 1
    for sheet in SHEETS:
        rows = by_sheet.get(sheet, [])
        tally = defaultdict(int)
        priors = 0
        for row in rows:
            tally[row["status"]] += 1
            if prior(row) == "YES":
                priors += 1
        ms = tally.get("ok", 0)
        nts = tally.get("not_to_standard", 0)
        ni = tally.get("not_installed", 0)
        sk = tally.get("skipped", 0)
        pend = tally.get("pending", 0)
        total = ms + nts + ni + sk + pend
        vals = [sheet, ms, nts, ni, sk, pend, total, priors]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = CELL_FONT
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            if ci == 7 and v != 508:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")  # flag if not 508
        r += 1

    note = ws.cell(r + 1, 1,
                   "Cross-foot check: every Total must equal 508 (the active-template "
                   "universe). Any cell flagged red breaks the audit -- investigate before "
                   "relying on the cert claim. A clean 248/252 shows 502 MS + 6 SKIP on the "
                   "C2 (top) inspections, 0 NI, 0 PENDING, 0 open defects.")
    note.font = Font(name="Aptos Narrow", size=9, italic=True, color=DARK)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=8)
    note.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    by_sheet = load()
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in SHEETS:
        ws = wb.create_sheet(sheet)
        n = write_audit_sheet(ws, by_sheet.get(sheet, []))
        print(f"{sheet}: {len(by_sheet.get(sheet, []))} item rows")
    recon = wb.create_sheet("Reconciliation")
    write_recon(recon, by_sheet)
    wb.save(OUT)
    print("saved", OUT)


if __name__ == "__main__":
    main()
