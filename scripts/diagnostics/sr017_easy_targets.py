"""
sr017_easy_targets.py

Builds the SR-017 Easy Targets xlsx (Sheets 1+2 only in this pass):
  Sheet 1 — Zone Summary  (per block / floor / cycle aggregation)
  Sheet 2 — Easy Targets  (89 pending units ranked ascending by cohort)

Saves to /tmp/sr017_easy_targets.xlsx then prints base64 between markers.

Retrieval on MacBook (one paste, between markers):
  pbpaste | base64 -d > ~/Desktop/sr017_easy_targets.xlsx
"""
import sqlite3
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TENANT_ID = 'MONOGRAPH'
BATCH_ID = '3237ea51'  # SR-017

# ----- Reuse v2 projection logic (preloads + project function) -----
conn = sqlite3.connect('/var/data/inspections.db')
conn.row_factory = sqlite3.Row
cur = conn.cursor()

cycle_meta = {}
for r in cur.execute(
        "SELECT id, cycle_number, block, floor FROM inspection_cycle WHERE tenant_id=?",
        (TENANT_ID,)).fetchall():
    cycle_meta[r['id']] = dict(r)

all_templates = {}
for r in cur.execute(
        "SELECT id, parent_item_id, floor_condition, active FROM item_template WHERE tenant_id=?",
        (TENANT_ID,)).fetchall():
    all_templates[r['id']] = dict(r)

parent_to_children = {}
for tid, t in all_templates.items():
    if t['parent_item_id']:
        parent_to_children.setdefault(t['parent_item_id'], []).append(tid)
parents_with_kids = set(parent_to_children.keys())

prev_items_by_unit_cycle = {}
for r in cur.execute(
        "SELECT i.unit_id, i.cycle_number, ii.item_template_id, ii.status "
        "FROM inspection_item ii JOIN inspection i ON ii.inspection_id = i.id "
        "WHERE i.tenant_id = ?", (TENANT_ID,)).fetchall():
    prev_items_by_unit_cycle.setdefault((r['unit_id'], r['cycle_number']), {})[r['item_template_id']] = r['status']

defects_by_unit = {}
for r in cur.execute(
        "SELECT unit_id, item_template_id, raised_cycle_id, raised_cycle_number, status, cleared_cycle_number "
        "FROM defect WHERE tenant_id = ?", (TENANT_ID,)).fetchall():
    defects_by_unit.setdefault(r['unit_id'], []).append(dict(r))

latents_by_unit = {}
for r in cur.execute(
        "SELECT unit_id, rectified_at, rectified_at_cycle_number "
        "FROM latent_area_note WHERE tenant_id = ?", (TENANT_ID,)).fetchall():
    latents_by_unit.setdefault(r['unit_id'], []).append(dict(r))

excl_by_list = {}
for r in cur.execute(
        "SELECT exclusion_list_id, item_template_id FROM exclusion_list_item").fetchall():
    excl_by_list.setdefault(r['exclusion_list_id'], set()).add(r['item_template_id'])

excl_by_cycle = {}
for r in cur.execute(
        "SELECT cycle_id, item_template_id FROM cycle_excluded_item WHERE tenant_id=?",
        (TENANT_ID,)).fetchall():
    excl_by_cycle.setdefault(r['cycle_id'], set()).add(r['item_template_id'])

def project(unit_id, cycle_id, excl_list_id, unit_floor, cycle_number):
    if excl_list_id and excl_list_id in excl_by_list:
        current_exclusions = excl_by_list[excl_list_id]
    elif cycle_id and cycle_id in excl_by_cycle:
        current_exclusions = excl_by_cycle[cycle_id]
    else:
        current_exclusions = set()

    prev_item_map = {}
    if cycle_number > 1:
        prev_item_map = prev_items_by_unit_cycle.get((unit_id, cycle_number - 1), {})

    prior_defect_set = set()
    for d in defects_by_unit.get(unit_id, []):
        if d['status'] == 'open' and d['raised_cycle_id'] != cycle_id:
            prior_defect_set.add(d['item_template_id'])

    c2_status, hpd = {}, {}
    for tid, t in all_templates.items():
        if t['active'] != 1:
            continue
        if tid in current_exclusions:
            s = 'skipped'
        elif t['floor_condition'] == 'ground_only' and unit_floor > 0:
            s = 'skipped'
        elif cycle_number > 1 and prev_item_map:
            prev_s = prev_item_map.get(tid)
            if not prev_s or prev_s == 'pending':
                s = 'ok'
            elif prev_s == 'ok':
                s = 'ok'
            elif prev_s in ('not_to_standard', 'not_installed'):
                s = 'pending'
            elif prev_s == 'skipped':
                s = 'pending'
            else:
                s = 'pending'
        else:
            prev_s = prev_item_map.get(tid)
            s = prev_s if prev_s else 'pending'
        c2_status[tid] = s
        hpd[tid] = 1 if tid in prior_defect_set else 0

    if cycle_number > 1:
        # Rule 2
        for tid in list(c2_status.keys()):
            t = all_templates[tid]
            if t['parent_item_id'] is not None or tid not in parents_with_kids:
                continue
            if c2_status[tid] != 'pending' or hpd[tid] != 0:
                continue
            kids = [c for c in parent_to_children.get(tid, []) if c in c2_status]
            if all(c2_status[c] == 'skipped' for c in kids):
                c2_status[tid] = 'ok'
        # Rule 3
        for tid in list(c2_status.keys()):
            t = all_templates[tid]
            if t['parent_item_id'] is None:
                continue
            if c2_status[tid] != 'pending' or hpd[tid] != 0:
                continue
            pid = t['parent_item_id']
            pt = all_templates.get(pid)
            if not pt or pt['parent_item_id'] is not None:
                continue
            if c2_status.get(pid) == 'pending' and hpd.get(pid, 0) == 0:
                c2_status[tid] = 'ok'
        # Rule 4
        for tid in list(c2_status.keys()):
            t = all_templates[tid]
            if t['parent_item_id'] is not None or tid not in parents_with_kids:
                continue
            if c2_status[tid] != 'pending' or hpd[tid] != 0:
                continue
            c2_status[tid] = 'ok'

    items = sum(1 for tid, s in c2_status.items() if s == 'pending' and hpd[tid] == 0)

    d_count = 0
    for d in defects_by_unit.get(unit_id, []):
        if d['raised_cycle_number'] is None or d['raised_cycle_number'] >= cycle_number:
            continue
        if d['status'] == 'open':
            d_count += 1
        elif d['status'] == 'cleared' and d['cleared_cycle_number'] == cycle_number:
            d_count += 1

    l_count = 0
    for l in latents_by_unit.get(unit_id, []):
        if l['rectified_at'] is None:
            l_count += 1
        elif l['rectified_at_cycle_number'] == cycle_number:
            l_count += 1

    return d_count, l_count, items

# ----- Compute pending units -----
results = []
for r in cur.execute("""
    SELECT bu.unit_id, bu.cycle_id, bu.exclusion_list_id, bu.status as bu_status,
           u.unit_number, u.floor, u.block
    FROM batch_unit bu JOIN unit u ON bu.unit_id = u.id
    WHERE bu.batch_id=? AND bu.tenant_id=? AND (bu.removed_at IS NULL OR bu.status != 'removed')
""", (BATCH_ID, TENANT_ID)).fetchall():
    try:
        floor = int(r['floor']) if r['floor'] is not None else 0
    except (ValueError, TypeError):
        floor = 0
    cm = cycle_meta.get(r['cycle_id'])
    cycle_number = cm['cycle_number'] if cm else 2
    d, l, items = project(r['unit_id'], r['cycle_id'], r['exclusion_list_id'], floor, cycle_number)
    results.append({
        'unit_number': r['unit_number'],
        'block': r['block'] or '',
        'floor': floor,
        'bu_status': r['bu_status'] or '',
        'cycle_number': cycle_number,
        'defects': d, 'latents': l, 'items': items, 'total': d + l + items,
    })

pending = [r for r in results if r['bu_status'] == 'pending']
pending.sort(key=lambda x: (x['total'], x['unit_number']))
print(f"# Pending: {len(pending)} units, total cohort = {sum(r['total'] for r in pending)}")

# ----- Build xlsx -----
wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill(start_color='F5F3EE', end_color='F5F3EE', fill_type='solid')
HEADER_FONT = Font(bold=True, size=11, color='2C2C2C')
THIN = Side(border_style='thin', color='D5D2CC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER

def style_cell(cell, align='center'):
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER

# === Sheet 1: Zone Summary ===
ws1 = wb.create_sheet('Zone Summary', 0)
ws1['A1'] = 'SR-017 Zone Summary'
ws1['A1'].font = Font(bold=True, size=14)
ws1.merge_cells('A1:I1')
ws1['A2'] = 'Pending units only — aggregated by Block × Floor × Cycle'
ws1['A2'].font = Font(italic=True, size=10, color='666666')
ws1.merge_cells('A2:I2')

headers1 = ['Block', 'Floor', 'Cycle', 'Units', 'Defects', 'Latents', 'New Items', 'Total Cohort', 'Avg / Unit']
for col, h in enumerate(headers1, 1):
    style_header(ws1.cell(row=4, column=col, value=h))

zones = {}
for r in pending:
    key = (r['block'], r['floor'], r['cycle_number'])
    z = zones.setdefault(key, {'units': 0, 'defects': 0, 'latents': 0, 'items': 0, 'total': 0})
    z['units'] += 1
    z['defects'] += r['defects']
    z['latents'] += r['latents']
    z['items'] += r['items']
    z['total'] += r['total']

row_idx = 5
for key in sorted(zones.keys()):
    block, floor, cyc = key
    z = zones[key]
    vals = [block, floor, f'C{cyc}', z['units'], z['defects'], z['latents'],
            z['items'], z['total'], round(z['total'] / z['units'], 1)]
    for col, v in enumerate(vals, 1):
        style_cell(ws1.cell(row=row_idx, column=col, value=v))
    row_idx += 1

# Totals row
total_units = sum(z['units'] for z in zones.values())
total_def = sum(z['defects'] for z in zones.values())
total_lat = sum(z['latents'] for z in zones.values())
total_items = sum(z['items'] for z in zones.values())
total_coh = sum(z['total'] for z in zones.values())
totals_row = ['TOTAL', '', '', total_units, total_def, total_lat, total_items, total_coh,
              round(total_coh / total_units, 1) if total_units else 0]
for col, v in enumerate(totals_row, 1):
    c = ws1.cell(row=row_idx, column=col, value=v)
    c.font = Font(bold=True)
    style_cell(c)
    c.fill = HEADER_FILL

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 8
ws1.column_dimensions['C'].width = 8
for col_letter in ['D','E','F','G','H','I']:
    ws1.column_dimensions[col_letter].width = 13
ws1.freeze_panes = 'A5'

# === Sheet 2: Easy Targets ===
ws2 = wb.create_sheet('Easy Targets', 1)
ws2['A1'] = 'SR-017 Easy Targets'
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:I1')
ws2['A2'] = f'{len(pending)} pending units — ranked ascending by total cohort'
ws2['A2'].font = Font(italic=True, size=10, color='666666')
ws2.merge_cells('A2:I2')

headers2 = ['Rank', 'Unit', 'Block', 'Floor', 'Cycle', 'Defects', 'Latents', 'New Items', 'Total Cohort']
for col, h in enumerate(headers2, 1):
    style_header(ws2.cell(row=4, column=col, value=h))

for i, r in enumerate(pending, start=1):
    row_idx = 4 + i
    vals = [i, r['unit_number'], r['block'], r['floor'], f"C{r['cycle_number']}",
            r['defects'], r['latents'], r['items'], r['total']]
    for col, v in enumerate(vals, 1):
        style_cell(ws2.cell(row=row_idx, column=col, value=v))

ws2.column_dimensions['A'].width = 7
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 8
for col_letter in ['F','G','H','I']:
    ws2.column_dimensions[col_letter].width = 13
ws2.freeze_panes = 'A5'
ws2.auto_filter.ref = f'A4:I{4 + len(pending)}'

# Save
xlsx_path = '/tmp/sr017_easy_targets.xlsx'
wb.save(xlsx_path)
print(f"# Saved {xlsx_path}")

# Base64
with open(xlsx_path, 'rb') as f:
    b64 = base64.b64encode(f.read()).decode('ascii')
print(f"# File size (base64): {len(b64)} chars\n")
print('=== BEGIN_BASE64 ===')
print(b64)
print('=== END_BASE64 ===')

conn.close()
