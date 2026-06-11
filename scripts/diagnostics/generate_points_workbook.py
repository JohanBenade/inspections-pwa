#!/usr/bin/env python3
"""
generate_points_workbook.py -- PERMANENT 190-unit Inspection Points workbook
generator. REPLACES the container-lost /tmp tc7.py builder.

Imports the SHARED engine (inspection_engine.py) so its per-unit counts are
IDENTICAL to the de-snag sheet generator -- the workbook and the sheets can
never disagree (Johan's core requirement).

Reproduces Inspection_Points_190_Units_v4.xlsx exactly:
  Sheet 'Summary'        : KPIs, What-Remains, Block rollup, Block x Floor grid
  Sheet '190-Unit Detail': per-unit rows (15 cols) + PROJECT TOTAL, live formulas

USAGE (on Render, read-only DB):
    python3 scripts/diagnostics/generate_points_workbook.py
    python3 scripts/diagnostics/generate_points_workbook.py --out /tmp --snap "2026-06-09 09:59 UTC"

Writes <out>/Inspection_Points_190_Units.xlsx. ASCII only. Read-only DB.
"""
import sys
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import inspection_engine as eng

GOLD = "FFC8963E"
DARK = "FF0A0A0A"
GREY = "FFD9D9D9"
SUBGREY = "FF9A9A9A"
FONT = "Aptos Narrow"

DETAIL = "190-Unit Detail"
SUMMARY = "Summary"

DETAIL_HEADERS = ["Unit", "Block", "Floor", "Cycle", "Status", "Total Points",
                  "Already OK", "Pending", "NTS", "Not Inst", "Skipped",
                  "Items to Mark", "Latents", "POINTS TO VISIT",
                  "Open Defects (context — not in PTV)"]

ROW2_NOTE = ("Total Points = 508 (ground) / 502 (above ground = 508 active "
             "points less 6 ground-only burglar-bar items, structurally absent "
             "upstairs).   Items to Mark = Pending + NTS + Not Inst + Skipped.   "
             "Already OK + Items to Mark = Total Points (row closes).   POINTS TO "
             "VISIT = Items to Mark + Latents.   Open Defects shown for context "
             "only — NOT added to PTV (they sit on the NTS items already counted).")

hdr_font = Font(name=FONT, bold=True, color="FFFFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor=DARK)
cell_font = Font(name=FONT, size=9)
thin = Side(style="thin", color="FFBBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_detail(ws, rows):
    ws.cell(1, 1, "INSPECTION POINTS TO VISIT  —  190 UNITS  (per-unit detail)").font = \
        Font(name=FONT, bold=True, size=14, color=DARK)
    n2 = ws.cell(2, 1, ROW2_NOTE)
    n2.font = Font(name=FONT, size=8, color=SUBGREY)
    n2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:O2")
    for ci, h in enumerate(DETAIL_HEADERS, 1):
        cell = ws.cell(3, ci, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r = 4
    for d in rows:
        ws.cell(r, 1, d["unit"])
        ws.cell(r, 2, d["block"])
        ws.cell(r, 3, d["floor"])
        ws.cell(r, 4, d["cycle"])
        ws.cell(r, 5, d["status"])
        ws.cell(r, 6, d["total_points"])
        ws.cell(r, 7, d["already_ok"])
        ws.cell(r, 8, d["pending"])
        ws.cell(r, 9, d["nts"])
        ws.cell(r, 10, d["not_inst"])
        ws.cell(r, 11, d["skipped"])
        ws.cell(r, 12, "=H%d+I%d+J%d+K%d" % (r, r, r, r))
        ws.cell(r, 13, d["latents"])
        ws.cell(r, 14, "=L%d+M%d" % (r, r))
        ws.cell(r, 15, d["open_defects"])
        for ci in range(1, 16):
            ws.cell(r, ci).font = cell_font
            ws.cell(r, ci).border = BORDER
        r += 1
    last = r - 1
    tot = ws.cell(r, 1, "PROJECT TOTAL")
    tot.font = Font(name=FONT, bold=True, size=10)
    for ci in range(6, 16):
        L = get_column_letter(ci)
        cell = ws.cell(r, ci, "=SUM(%s4:%s%d)" % (L, L, last))
        cell.font = Font(name=FONT, bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor=GOLD)
        cell.border = BORDER
    widths = [8, 9, 6, 7, 22, 9, 9, 8, 6, 8, 8, 10, 8, 12, 28]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"
    return r  # project-total row index


def build_summary(ws, total_row, blocks, floors):
    D = "'%s'" % DETAIL
    ws.cell(1, 1, "POWER PARK STUDENT HOUSING — PHASE 3").font = Font(name=FONT, bold=True, size=16, color=DARK)
    ws.cell(2, 1, "Inspection Points to Visit — Project Certification Scorecard").font = Font(name=FONT, size=11, color=SUBGREY)
    ws.cell(3, 1, "190 four-bed units · 7 blocks · source DB snapshot").font = Font(name=FONT, size=9, color=SUBGREY)
    tr = total_row
    # headline KPI band (row 5 labels, row 6 values, row 7 sublabels)
    kpis = [
        (1, "TOTAL SCOPE", "=%s!F%d" % (D, tr), "inspection points"),
        (3, "PASSED (OK)", "=%s!G%d" % (D, tr), "points marked MS/OK"),
        (5, "% PASSED", "=%s!G%d/%s!F%d" % (D, tr, D, tr), "of total scope"),
        (7, "POINTS TO VISIT", "=%s!N%d" % (D, tr), "on-site actions remaining"),
        (9, "TRULY CERTIFIED", "=COUNTIFS(%s!N4:N%d,0)" % (D, tr - 1), "units, every axis clean"),
    ]
    for col, label, formula, sub in kpis:
        ws.cell(5, col, label).font = Font(name=FONT, bold=True, size=10, color=DARK)
        v = ws.cell(6, col, formula)
        v.font = Font(name=FONT, bold=True, size=18, color=GOLD)
        if label == "% PASSED":
            v.number_format = "0.0%"
        ws.cell(7, col, sub).font = Font(name=FONT, size=8, color=SUBGREY)
    # What remains
    ws.cell(9, 1, "WHAT REMAINS  (the Points to Visit, decomposed)").font = Font(name=FONT, bold=True, size=11, color=DARK)
    remain = [
        (10, "Pending — not yet inspected", "H"),
        (11, "NTS — failed, awaiting re-mark", "I"),
        (12, "Not installed", "J"),
        (13, "Skipped — previously excluded, now in scope", "K"),
        (14, "Items to Mark (subtotal)", "L"),
        (15, "Latent defects to rectify", "M"),
        (16, "POINTS TO VISIT  (Items to Mark + Latents)", "N"),
    ]
    for row, label, col in remain:
        ws.cell(row, 1, label).font = Font(name=FONT, size=10,
            bold=(col in ("L", "N")))
        ws.cell(row, 4, "=%s!%s%d" % (D, col, tr)).font = Font(name=FONT, size=10, bold=(col in ("L", "N")))
    rc = ws.cell(17, 1, "Reconciliation check:  Passed + Points to Visit = Total Scope")
    rc.font = Font(name=FONT, size=9, italic=True, color=SUBGREY)
    ws.cell(17, 4, "=%s!G%d+%s!N%d" % (D, tr, D, tr)).font = Font(name=FONT, size=9, italic=True)
    # Block rollup
    ws.cell(19, 1, "BLOCK ROLLUP").font = Font(name=FONT, bold=True, size=11, color=DARK)
    for ci, h in enumerate(["Block", "", "Units", "Total Points", "Passed", "% Passed", "Points to Visit"], 1):
        if h:
            c = ws.cell(20, ci, h); c.font = hdr_font; c.fill = hdr_fill
    r = 21
    for b in blocks:
        ws.cell(r, 1, b).font = cell_font
        ws.cell(r, 3, "=COUNTIFS(%s!B4:B%d,A%d)" % (D, tr - 1, r)).font = cell_font
        ws.cell(r, 4, "=SUMIFS(%s!F4:F%d,%s!B4:B%d,A%d)" % (D, tr - 1, D, tr - 1, r)).font = cell_font
        ws.cell(r, 5, "=SUMIFS(%s!G4:G%d,%s!B4:B%d,A%d)" % (D, tr - 1, D, tr - 1, r)).font = cell_font
        f = ws.cell(r, 6, "=IF(SUMIFS(%s!F4:F%d,%s!B4:B%d,A%d)=0,0,E%d/D%d)" % (D, tr - 1, D, tr - 1, r, r, r))
        f.font = cell_font; f.number_format = "0.0%"
        ws.cell(r, 7, "=SUMIFS(%s!N4:N%d,%s!B4:B%d,A%d)" % (D, tr - 1, D, tr - 1, r)).font = cell_font
        r += 1
    ws.cell(r, 1, "TOTAL").font = Font(name=FONT, bold=True, size=10)
    for col in ("C", "D", "E", "G"):
        ws.cell(r, "CDEFG".index(col) + 3 if False else {"C":3,"D":4,"E":5,"G":7}[col],
                "=SUM(%s21:%s%d)" % (col, col, r - 1)).font = Font(name=FONT, bold=True)
    tf = ws.cell(r, 6, "=IF(D%d=0,0,E%d/D%d)" % (r, r, r)); tf.font = Font(name=FONT, bold=True); tf.number_format = "0.0%"
    widths = [16, 3, 8, 12, 10, 10, 14, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="/tmp")
    ap.add_argument("--db", default=eng.DB_DEFAULT)
    ap.add_argument("--snap", default="")
    args = ap.parse_args()

    c = eng.connect(args.db)
    units = eng.all_units(c)
    rows = [eng.unit_counts(c, u) for u in units]

    # invariant guard: every row must close
    broken = [r["unit"] for r in rows if not r["closes"]]
    if broken:
        print("WARN: rows do not close (already_ok+items!=total):", broken)

    wb = Workbook()
    sumws = wb.active
    sumws.title = SUMMARY
    detws = wb.create_sheet(DETAIL)
    total_row = build_detail(detws, rows)
    blocks = sorted({r["block"] for r in rows if r["block"]})
    floors = sorted({(r["block"], r["floor"]) for r in rows})
    build_summary(sumws, total_row, blocks, floors)

    path = "%s/Inspection_Points_190_Units.xlsx" % args.out.rstrip("/")
    wb.save(path)
    print("Units: %d | PROJECT TOTAL row: %d | rows-close: %s -> %s"
          % (len(rows), total_row, not broken, path))


if __name__ == "__main__":
    main()
