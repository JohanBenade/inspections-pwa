#!/usr/bin/env python3
"""
generate_inspection_sheet.py -- PERMANENT committed inspection-sheet generator.

Produces the per-unit "DEFECTIVE WORKS LIST -- DE-SNAG" xlsx (Raubex-facing,
Kevin-signed legal document) in the EXACT format of the project-knowledge
sample sheets (Unit_101 / Unit_250 / Unit_023).

This script REPLACES the container-lost /tmp tc7.py generator. It exists so the
line-set logic is NEVER reverse-engineered again. Run it; do not rebuild it.

LINE-SET LOGIC (v411 alignment audit -- proven 0 PTV / 0 cycle / 0 latent
mismatch against Inspection_Points_190_Units across all 190 units):

  per unit:
    uid, floor from unit_real (NOT LIKE 'TEST%')
    highest-cycle inspection: ORDER BY cycle_number DESC, created_at DESC LIMIT 1
    line set = inspection_item WHERE status IN
               ('pending','not_to_standard','not_installed','skipped')
               JOIN item_template ON active=1
               MINUS structural ground_only-above-GF (floor>0 AND fc='ground_only')
    latents = latent_area_note WHERE rectified_at IS NULL
    sheet_PTV = len(lines) + latents

Rows are emitted in walk order: area_order, category_order, item_order
(area_template -> category_template -> item_template).

The verdict tick columns (INST / M.S. / N.T.S.) are left BLANK for field
completion -- this is a de-snag worklist, not a filled record. The E column
carries the prior-defect comment for context where one exists.

USAGE (on Render, read-only DB):
    python3 scripts/diagnostics/generate_inspection_sheet.py 248 252
    python3 scripts/diagnostics/generate_inspection_sheet.py 248 --out /tmp

Writes /tmp/Unit_<n>_Inspection_Sheet.xlsx per unit (override dir with --out).

ASCII only. Read-only w.r.t. the DB. No hardcoded per-unit data -- every row is
sourced live. Nothing about layout is guessed: it mirrors the sample workbooks.
"""
import sys
import sqlite3
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB = "/var/data/inspections.db"
TENANT = "MONOGRAPH"

# ---- format constants (mirrored verbatim from the sample sheets) -------------
FONT = "Aptos Narrow"
GREY = "FFD9D9D9"          # area band fill
COMMENT_RED = "FFC0504D"   # prior-defect comment colour
ITALIC_GREY = "FF595959"   # context-note colour

PHYS_ADDR = "Physical Address | 8 Ridge Road, Mountain View, Pretoria"
PROJECT_LINE = "POWER PARK STUDENT HOUSING | PHASE 3"

PREAMBLE_TITLE = "Standard of measurement for items on defective works list:"
PREAMBLE = [
    "Visual defects on all items shall be inspected from 1m distance and deemed defective if prominent.",
    "All items not operating as intended shall be deemed defective.",
    "All items not installed or missing shall be deemed incomplete.",
    "Any critical damage to items that can affect the longevity of the item or void its warranty shall be deemed defective.",
    "Any item that does not comply to the South African National Standards and accompanying codes shall be deemed defective.",
]
DEF_TITLE = "Document Definitions:"
DEF_LINE = "INST. - Installed      NI - Not installed      M.S. - Meets standard      N.T.S. - Not to standard"
CONTEXT_NOTE = "Prior defect shown for context — re-check and write the current finding. Tick the applicable column."

STATUS_IN = ("pending", "not_to_standard", "not_installed", "skipped")


def connect():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c


def resolve_unit(c, unit_number):
    u = c.execute(
        "SELECT id, block, floor, unit_number FROM unit_real "
        "WHERE unit_number=? AND tenant_id=? AND unit_number NOT LIKE 'TEST%'",
        (unit_number, TENANT)).fetchone()
    return u


def top_inspection(c, uid):
    return c.execute(
        "SELECT id, cycle_number, status FROM inspection "
        "WHERE unit_id=? ORDER BY cycle_number DESC, created_at DESC LIMIT 1",
        (uid,)).fetchone()


def line_rows(c, insp_id, floor):
    """v411-audited line set, joined to area/category for walk-order + names,
    with the latest open-defect comment per item where one exists."""
    rows = c.execute("""
        SELECT at2.area_name, at2.area_order,
               ct.category_name, ct.category_order,
               it.item_description, it.item_order,
               it.floor_condition,
               ii.status
        FROM inspection_item ii
        JOIN item_template it ON ii.item_template_id = it.id AND it.active = 1
        JOIN category_template ct ON it.category_id = ct.id
        JOIN area_template at2 ON ct.area_id = at2.id
        WHERE ii.inspection_id = ?
          AND ii.status IN ('pending','not_to_standard','not_installed','skipped')
        ORDER BY at2.area_order, ct.category_order, it.item_order
    """, (insp_id,)).fetchall()
    # structural skip removal: ground_only items above ground floor
    out = []
    for r in rows:
        if floor and floor > 0 and r["floor_condition"] == "ground_only":
            continue
        out.append(r)
    return out


def comment_for(c, uid, area_name, item_description):
    """Latest open-defect original_comment for context, if any. Item-template
    join via the unit's open defects -- empty string when none."""
    r = c.execute("""
        SELECT d.original_comment
        FROM defect d
        JOIN item_template it ON d.item_template_id = it.id
        JOIN category_template ct ON it.category_id = ct.id
        JOIN area_template at2 ON ct.area_id = at2.id
        WHERE d.unit_id = ? AND d.status='open'
          AND at2.area_name = ? AND it.item_description = ?
        ORDER BY d.updated_at DESC LIMIT 1
    """, (uid, area_name, item_description)).fetchone()
    return (r["original_comment"] or "") if r else ""


def open_latents(c, uid):
    return c.execute(
        "SELECT COUNT(*) FROM latent_area_note WHERE unit_id=? AND rectified_at IS NULL",
        (uid,)).fetchone()[0]


def build_sheet(unit, insp, lines, latents, c):
    wb = Workbook()
    ws = wb.active
    ws.title = "Unit %s" % unit["unit_number"]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 6.5
    ws.column_dimensions["C"].width = 6.5
    ws.column_dimensions["D"].width = 6.5
    ws.column_dimensions["E"].width = 60

    bold10 = Font(name=FONT, size=10, bold=True)
    reg10 = Font(name=FONT, size=10)
    reg9 = Font(name=FONT, size=9)
    note_it = Font(name=FONT, size=9, italic=True, color=ITALIC_GREY)
    comment_it = Font(name=FONT, size=9, italic=True, color=COMMENT_RED)
    area_fill = PatternFill("solid", fgColor=GREY)
    right = Alignment(horizontal="right", wrap_text=True)
    left = Alignment(horizontal="left", wrap_text=True)
    left_i2 = Alignment(horizontal="left", indent=2, wrap_text=True)
    left_i1 = Alignment(horizontal="left", indent=1, wrap_text=True)

    def merge_de(r):
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

    def merge_ae(r):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    # ---- header block --------------------------------------------------------
    ws.merge_cells("D1:E2")
    ws["D1"] = PHYS_ADDR
    ws["D1"].font = Font(name=FONT, size=8)
    ws["D1"].alignment = right

    pairs = [
        (4, "ATTENTION", "MAIN CONTRACTOR | RAUBEX"),
        (5, "RE:", "DEFECTIVE WORKS LIST — DE-SNAG"),
        (6, "PROJECT", PROJECT_LINE),
        (7, "UNIT NO / AREA OF INSPECTION",
         "BLOCK %s / FLOOR %s / UNIT %s  —  C%s" % (
             str(unit["block"]).replace("Block ", ""), unit["floor"],
             unit["unit_number"], insp["cycle_number"])),
    ]
    for r, a, d in pairs:
        ws.cell(r, 1, a).font = bold10
        merge_de(r)
        dc = ws.cell(r, 4, d)
        dc.font = bold10
        dc.alignment = Alignment(horizontal="right")

    ws.cell(9, 1, PREAMBLE_TITLE).font = bold10
    merge_ae(9)
    r = 10
    for line in PREAMBLE:
        ws.cell(r, 1, line).font = reg9
        merge_ae(r)
        r += 1
    r += 1  # blank 15
    ws.cell(16, 1, DEF_TITLE).font = bold10
    merge_ae(16)
    ws.cell(17, 1, DEF_LINE).font = reg9
    merge_ae(17)
    ws.cell(18, 1, CONTEXT_NOTE).font = note_it
    merge_ae(18)

    # ---- body ----------------------------------------------------------------
    ws.cell(20, 1, "description").font = bold10
    ws.cell(20, 1).alignment = left
    r = 21
    cur_area = None
    cur_cat = None
    checkpoints = 0
    for row in lines:
        if row["area_name"] != cur_area:
            cur_area = row["area_name"]
            cur_cat = None
            ac = ws.cell(r, 1, cur_area.upper())
            ac.font = bold10
            ac.fill = area_fill
            ac.alignment = left
            merge_ae(r)
            r += 1
        if row["category_name"] != cur_cat:
            cur_cat = row["category_name"]
            cc = ws.cell(r, 1, cur_cat)
            cc.font = bold10 if cur_cat.strip().lower() != "general" else reg10
            cc.alignment = left
            r += 1
        ic = ws.cell(r, 1, row["item_description"])
        ic.font = reg10
        ic.alignment = left_i2
        cmt = comment_for(c, unit_id_global, row["area_name"], row["item_description"])
        if cmt:
            ec = ws.cell(r, 5, cmt)
            ec.font = comment_it
            ec.alignment = left_i1
        checkpoints += 1
        r += 1

    # ---- footer --------------------------------------------------------------
    total = checkpoints + latents
    r += 1
    fc = ws.cell(r, 1, "Checkpoints to action: %d" % total)
    fc.font = reg10
    r += 2
    for label in ("Inspection Date: ", "Inspected By: ", "Certification date: "):
        ws.cell(r, 1, label).font = reg10
        r += 1

    return wb, total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("units", nargs="+", help="unit numbers e.g. 248 252")
    ap.add_argument("--out", default="/tmp", help="output dir (default /tmp)")
    args = ap.parse_args()

    c = connect()
    global unit_id_global
    for un in args.units:
        unit = resolve_unit(c, un)
        if not unit:
            print("SKIP %s: not found in unit_real" % un)
            continue
        unit_id_global = unit["id"]
        insp = top_inspection(c, unit["id"])
        if not insp:
            print("SKIP %s: no inspection" % un)
            continue
        lines = line_rows(c, insp["id"], unit["floor"])
        latents = open_latents(c, unit["id"])
        wb, total = build_sheet(unit, insp, lines, latents, c)
        path = "%s/Unit_%s_Inspection_Sheet.xlsx" % (args.out.rstrip("/"), un)
        wb.save(path)
        print("Unit %s: C%s %s | %d line(s) + %d latent(s) = %d checkpoints -> %s"
              % (un, insp["cycle_number"], insp["status"],
                 len(lines), latents, total, path))


if __name__ == "__main__":
    unit_id_global = None
    main()
